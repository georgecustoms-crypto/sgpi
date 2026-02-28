import flet as ft
import psycopg2
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from urllib.parse import urlparse

# ================= CONFIG =================
APP_FOLDER = os.path.join(os.getcwd(), "data")
os.makedirs(APP_FOLDER, exist_ok=True)
BACKUP_FOLDER = os.path.join(APP_FOLDER, "backups")
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# ================= BANCO =================
DATABASE_URL = "postgres://postgres:26ge453*t28@db.hulqfzeslzusxhvdnnhu.supabase.co:5432/postgres"

def get_conn():
    result = urlparse(DATABASE_URL)
    return psycopg2.connect(
        dbname=result.path[1:],
        user=result.username,
        password=result.password,
        host=result.hostname,
        port=result.port
    )

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS salas (
            id SERIAL PRIMARY KEY,
            proprietario TEXT,
            andar TEXT,
            sala TEXT,
            empresa TEXT,
            tipo_escritorio TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            usuario TEXT UNIQUE,
            senha TEXT,
            nivel TEXT
        )
    """)
    cur.execute("SELECT * FROM usuarios WHERE usuario='admin'")
    if not cur.fetchone():
        cur.execute(
            "INSERT INTO usuarios (usuario, senha, nivel) VALUES (%s, %s, %s)",
            ("admin", "123", "admin")
        )
    conn.commit()
    cur.close()
    conn.close()

# ================= FUNÇÕES =================
def verificar_login(usuario, senha):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT nivel FROM usuarios WHERE usuario=%s AND senha=%s", (usuario, senha))
    resultado = cur.fetchone()
    cur.close()
    conn.close()
    return resultado

def listar_usuarios():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, usuario, nivel FROM usuarios")
    dados = cur.fetchall()
    cur.close()
    conn.close()
    return dados

def adicionar_usuario(usuario, senha, nivel):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO usuarios (usuario, senha, nivel) VALUES (%s, %s, %s)", (usuario, senha, nivel))
    conn.commit()
    cur.close()
    conn.close()

def excluir_usuario(uid):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=%s", (uid,))
    conn.commit()
    cur.close()
    conn.close()

def inserir_sala(proprietario, andar, sala, empresa, tipo):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO salas (proprietario, andar, sala, empresa, tipo_escritorio)
        VALUES (%s, %s, %s, %s, %s)
    """, (proprietario, andar, sala, empresa, tipo))
    conn.commit()
    cur.close()
    conn.close()

def buscar_salas(termo=""):
    conn = get_conn()
    cur = conn.cursor()
    like_term = f"%{termo}%"
    cur.execute("""
        SELECT proprietario, andar, sala, empresa, tipo_escritorio
        FROM salas
        WHERE empresa ILIKE %s OR sala ILIKE %s OR proprietario ILIKE %s
    """, (like_term, like_term, like_term))
    dados = cur.fetchall()
    cur.close()
    conn.close()
    return dados

def realizar_backup():
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(BACKUP_FOLDER, f"backup_{timestamp}.sql")
        os.system(f"pg_dump {DATABASE_URL} > {backup_file}")
        return backup_file
    except:
        return None

def importar_salas_excel(caminho):
    if not os.path.exists(caminho):
        return 0
    conn = get_conn()
    cur = conn.cursor()
    wb = load_workbook(caminho)
    sheet = wb.active
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cur.execute("""
            INSERT INTO salas (proprietario, andar, sala, empresa, tipo_escritorio)
            VALUES (%s, %s, %s, %s, %s)
        """, row)
        total += 1
    conn.commit()
    cur.close()
    conn.close()
    return total

# ================= APP =================
def main(page: ft.Page):
    page.title = "SGPI – Sistema de Gestão Predial Inteligente"
    page.theme_mode = ft.ThemeMode.LIGHT
    usuario_logado = {"nivel": None, "nome": ""}

    # -------- LOGIN --------
    def tela_login():
        page.clean()
        page.vertical_alignment = ft.MainAxisAlignment.CENTER
        page.horizontal_alignment = ft.MainAxisAlignment.CENTER

        usuario = ft.TextField(label="Usuário", prefix_icon="person", width=350)
        senha = ft.TextField(label="Senha", password=True, prefix_icon="lock", width=350)
        mensagem = ft.Text("")

        def entrar(e):
            resultado = verificar_login(usuario.value, senha.value)
            if resultado:
                usuario_logado["nivel"] = resultado[0]
                usuario_logado["nome"] = usuario.value
                dashboard()
            else:
                mensagem.value = "Usuário ou senha inválidos"
                mensagem.color = ft.colors.RED
                page.update()

        page.add(
            ft.Card(
                content=ft.Container(
                    padding=40,
                    width=450,
                    content=ft.Column([
                        ft.Text("Bem-vindo ao SGPI", size=24, weight="bold"),
                        ft.Text("Faça login para continuar"),
                        usuario,
                        senha,
                        ft.ElevatedButton("Entrar", icon="login", on_click=entrar),
                        mensagem
                    ], spacing=20, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
                )
            )
        )

    # -------- DASHBOARD --------
    def dashboard():
        page.clean()
        dialog = ft.AlertDialog(modal=True)
        page.dialog = dialog

        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Proprietário")),
                ft.DataColumn(ft.Text("Andar")),
                ft.DataColumn(ft.Text("Sala")),
                ft.DataColumn(ft.Text("Empresa")),
                ft.DataColumn(ft.Text("Tipo")),
            ],
            rows=[]
        )

        def atualizar(termo=""):
            tabela.rows.clear()
            for r in buscar_salas(termo):
                tabela.rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text(str(c))) for c in r]))
            page.update()

        campo_busca = ft.TextField(
            label="Buscar por Proprietário, Sala ou Empresa",
            prefix_icon="search",
            on_change=lambda e: atualizar(e.control.value)
        )

        # File picker
        file_picker = ft.FilePicker()
        page.overlay.append(file_picker)

        def arquivo_escolhido(e):
            if e.files:
                total = importar_salas_excel(e.files[0].path)
                atualizar()
                page.snack_bar = ft.SnackBar(ft.Text(f"{total} registros importados!"))
                page.snack_bar.open = True
                page.update()

        file_picker.on_result = arquivo_escolhido

        # Cadastro sala
        def abrir_cadastro(e):
            p = ft.TextField(label="Proprietário")
            a = ft.TextField(label="Andar")
            s = ft.TextField(label="Sala")
            em = ft.TextField(label="Empresa")
            t = ft.TextField(label="Tipo Escritório")

            def salvar(e):
                inserir_sala(p.value, a.value, s.value, em.value, t.value)
                dialog.open = False
                atualizar()
                page.update()

            def cancelar(e):
                dialog.open = False
                page.update()

            dialog.title = ft.Text("Cadastro de Sala", size=20, weight="bold")
            dialog.content = ft.Column([p, a, s, em, t])
            dialog.actions = [
                ft.ElevatedButton("Salvar", on_click=salvar),
                ft.TextButton("Cancelar", on_click=cancelar)
            ]
            dialog.open = True
            page.update()

        # Usuários
        tabela_usuarios = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Usuário")),
                ft.DataColumn(ft.Text("Nível")),
                ft.DataColumn(ft.Text("Excluir")),
            ],
            rows=[]
        )

        def atualizar_usuarios():
            tabela_usuarios.rows.clear()
            for uid, usuario, nivel in listar_usuarios():
                tabela_usuarios.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(ft.Text(usuario)),
                        ft.DataCell(ft.Text(nivel)),
                        ft.DataCell(ft.IconButton(
                            icon="delete",
                            on_click=lambda e, id=uid: deletar_usuario(id)
                        ))
                    ])
                )
            page.update()

        def deletar_usuario(uid):
            excluir_usuario(uid)
            atualizar_usuarios()

        def novo_usuario(e):
            u = ft.TextField(label="Usuário")
            s = ft.TextField(label="Senha")
            n = ft.Dropdown(label="Nível", options=[
                ft.dropdown.Option("admin"),
                ft.dropdown.Option("operador")
            ])

            def salvar(e):
                nivel_selecionado = n.value or "operador"
                adicionar_usuario(u.value, s.value, nivel_selecionado)
                dialog.open = False
                atualizar_usuarios()
                page.update()

            def cancelar(e):
                dialog.open = False
                page.update()

            dialog.title = ft.Text("Novo Usuário", size=20, weight="bold")
            dialog.content = ft.Column([u, s, n])
            dialog.actions = [
                ft.ElevatedButton("Salvar", on_click=salvar),
                ft.TextButton("Cancelar", on_click=cancelar)
            ]
            dialog.open = True
            page.update()

        atualizar()
        atualizar_usuarios()

        # Backup
        def backup_dialog(page):
            caminho = realizar_backup()
            if caminho:
                msg = f"Backup realizado com sucesso em:\n{caminho}"
            else:
                msg = "Erro ao realizar backup!"

            def fechar_backup(e):
                page.dialog.open = False
                page.update()

            page.dialog = ft.AlertDialog(
                title=ft.Text("Backup"),
                content=ft.Text(msg),
                actions=[ft.ElevatedButton("Fechar", on_click=fechar_backup)]
            )
            page.dialog.open = True
            page.update()

        # Tabs
        tabs = [
            ft.Tab(text="Consulta", content=ft.Column([campo_busca, tabela])),
            ft.Tab(text="Admin Salas", content=ft.Column([
                ft.ElevatedButton("Cadastrar Sala", on_click=abrir_cadastro),
                ft.ElevatedButton("Importar Excel", on_click=lambda e: file_picker.pick_files(allowed_extensions=["xlsx"])),
                ft.ElevatedButton("Gerar Backup", on_click=lambda e: backup_dialog(page))
            ]))
        ]

        if usuario_logado["nivel"] == "admin":
            tabs.append(
                ft.Tab(text="Usuários", content=ft.Column([
                    ft.ElevatedButton("Novo Usuário", on_click=novo_usuario),
                    tabela_usuarios
                ]))
            )

        page.add(ft.Tabs(tabs=tabs))

    tela_login()

# ================= MAIN =================
if __name__ == "__main__":
    init_db()
    ft.app(target=main, host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))